from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

DOCS = [
    ("Contract", "Signed Contract"),
    ("Contract", "Approved Proposal"),
    ("Contract", "Change Orders"),
    ("Scope", "Final SOW"),
    ("Scope", "System Design"),
    ("Scope", "Equipment List (BOM)"),
    ("Drawings", "Floor Plans"),
    ("Drawings", "Device Locations"),
    ("Drawings", "Cable Pathways"),
    ("Engineering", "RFIs"),
    ("Engineering", "ASIs"),
    ("Logistics", "Schedule"),
    ("Logistics", "Site Conditions"),
    ("Logistics", "Permits"),
    ("Client", "Contacts"),
    ("Client", "GC Info"),
]


def style_header(ws) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    ws.freeze_panes = "A2"


def set_column_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_workbook(output_path: Path) -> Path:
    wb = Workbook()

    # Dashboard sheet
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    headers_dash = [
        "Project Name",
        "Sales Rep",
        "PM",
        "Start Date",
        "Total Docs",
        "Completed Docs",
        "% Complete",
        "Kickoff Ready",
        "Status",
        "Notes",
    ]
    ws_dash.append(headers_dash)

    ws_dash["E2"] = '=COUNTIFS(\'Master Checklist\'!A:A,A2,\'Master Checklist\'!D:D,"Y")'
    ws_dash["F2"] = '=COUNTIFS(\'Master Checklist\'!A:A,A2,\'Master Checklist\'!G:G,"Y")'
    ws_dash["G2"] = '=IFERROR(F2/E2,0)'
    ws_dash["H2"] = '=IF(G2>=0.95,"YES","NO")'
    ws_dash["I2"] = '=IF(H2="YES","Ready","Hold")'
    ws_dash["G2"].number_format = "0%"

    style_header(ws_dash)
    set_column_widths(
        ws_dash,
        {
            "A": 28,
            "B": 16,
            "C": 16,
            "D": 14,
            "E": 12,
            "F": 14,
            "G": 12,
            "H": 14,
            "I": 12,
            "J": 42,
        },
    )

    # Master Checklist sheet
    ws_master = wb.create_sheet("Master Checklist")

    headers_master = [
        "Project Name",
        "Category",
        "Document",
        "Required (Y/N)",
        "Submitted (Y/N)",
        "PM Reviewed (Y/N)",
        "Complete (Auto)",
    ]
    ws_master.append(headers_master)

    start_row_master = 2
    for idx, (category, document) in enumerate(DOCS, start=start_row_master):
        ws_master.append(["", category, document, "Y", "", "", ""])
        ws_master[f"E{idx}"] = (
            f'=IF(COUNTIFS(\'Sales Input\'!$A:$A,$A{idx},\'Sales Input\'!$B:$B,$C{idx},'
            f'\'Sales Input\'!$C:$C,"Y")>0,"Y","")'
        )
        ws_master[f"G{idx}"] = f'=IF(AND(D{idx}="Y",E{idx}="Y",F{idx}="Y"),"Y","N")'

    style_header(ws_master)
    set_column_widths(
        ws_master,
        {"A": 28, "B": 16, "C": 32, "D": 16, "E": 16, "F": 16, "G": 16},
    )

    # Sales Input sheet
    ws_sales = wb.create_sheet("Sales Input")
    headers_sales = ["Project Name", "Document", "Complete (Y/N)", "Notes"]
    ws_sales.append(headers_sales)

    for _, document in DOCS:
        ws_sales.append(["", document, "", ""])

    style_header(ws_sales)
    set_column_widths(ws_sales, {"A": 28, "B": 32, "C": 16, "D": 42})

    # Kickoff ROM Summary sheet
    ws_kickoff = wb.create_sheet("Kickoff ROM Summary")
    headers_kickoff = [
        "Project Name",
        "Doc No (Auto Format)",
        "Risk Status",
        "Missing Required Count",
        "Narrative ROM Generated (Y/N)",
        "PDF Exported (Y/N)",
        "Assumptions Included (Y/N)",
        "Kickoff Decision",
        "Comments",
    ]
    ws_kickoff.append(headers_kickoff)

    ws_kickoff["H2"] = '=IF(AND(E2="Y",F2="Y",G2="Y",C2<>"Red"),"GO","HOLD")'

    style_header(ws_kickoff)
    set_column_widths(
        ws_kickoff,
        {
            "A": 28,
            "B": 32,
            "C": 14,
            "D": 20,
            "E": 24,
            "F": 18,
            "G": 22,
            "H": 16,
            "I": 42,
        },
    )

    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "HOW TO USE THIS TRACKER:",
        "",
        "1) SALES:",
        "- Fill out 'Sales Input' tab",
        "- Mark completed documents as Y",
        "",
        "2) PROJECT MANAGER:",
        "- Use 'Master Checklist' tab",
        "- PM marks 'PM Reviewed (Y/N)'",
        "- 'Submitted (Y/N)' auto-syncs from Sales Input",
        "",
        "3) DASHBOARD:",
        "- Enter project metadata in row 2+",
        "- Copy formulas in E2:I2 down for additional projects",
        "",
        "4) KICKOFF ROM SUMMARY:",
        "- Track risk, narrative generation, and PDF readiness",
        "- Kickoff decision auto-calculates in H column",
        "",
        "RULE:",
        "No documentation = No kickoff = No installation",
    ]

    for line in instructions:
        ws_inst.append([line])

    ws_inst.column_dimensions["A"].width = 90

    # Y/N dropdowns
    yn_master = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws_master.add_data_validation(yn_master)
    yn_master.add(f"D{start_row_master}:D{start_row_master + len(DOCS) - 1}")
    yn_master.add(f"F{start_row_master}:F{start_row_master + len(DOCS) - 1}")

    yn_sales = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws_sales.add_data_validation(yn_sales)
    yn_sales.add(f"C2:C{1 + len(DOCS)}")

    yn_kickoff = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws_kickoff.add_data_validation(yn_kickoff)
    yn_kickoff.add("E2:G500")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    default_output = Path(
        r"C:\Users\art\OneDrive\Documents\Morgan Sound ROM Builder\Kickoff_Integrated_Dashboard.xlsx"
    )
    saved = build_workbook(default_output)
    print(f"Saved tracker workbook: {saved}")


if __name__ == "__main__":
    main()
